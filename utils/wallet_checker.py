import os
import asyncio
import pandas as pd

from termcolor import cprint, colored
from prettytable import PrettyTable
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from modules.interfaces import SoftwareException
from utils.networks import EthereumRPC
from modules import Client
from dev import GeneralSettings
from config import ACCOUNT_NAMES, PRIVATE_KEYS, PROXIES


class TxChecker:
    @staticmethod
    async def get_nonce(client):
        url = 'https://claims.movementnetwork.xyz/api/get-nonce'

        headers = {
            'accept': '*/*',
            'accept-language': 'ru,en-US;q=0.9,en;q=0.8,ru-RU;q=0.7',
            'priority': 'u=1, i',
            'referer': 'https://claims.movementnetwork.xyz/',
            'sec-ch-ua': '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
        }

        response = await client.session.get(url, headers=headers)

        if response.status == 200:
            response_data = await response.json()
            return response_data['nonce']
        else:
            raise SoftwareException(f"Can`t get nonce, response: {await response.text()}")

    async def get_drop_amount(self, client):
        headers = {
            'accept': '*/*',
            'accept-language': 'ru,en-US;q=0.9,en;q=0.8,ru-RU;q=0.7',
            'priority': 'u=1, i',
            'referer': 'https://claims.movementnetwork.xyz/',
            'sec-ch-ua': '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
        }

        try:
            nonce = await self.get_nonce(client)

            url = 'https://claims.movementnetwork.xyz/api/claim/start'

            msg_to_sign = f'Please sign this message to confirm ownership. nonce: {nonce}'

            signature = await client.sign_message(msg_to_sign)

            json_data = {
                'address': f"{client.address}",
                'message': f'Please sign this message to confirm ownership. nonce: {nonce}',
                'signature': signature,
                'nonce': nonce,
            }

            response = await client.session.post(url, headers=headers, json=json_data)

            if response.status == 200:
                response_data = await response.json()
                if response_data.get('isEligible'):
                    move_drop, move_l2_drop = response_data['amount'], response_data['amountL2']
                    client.logger_msg(
                        *client.acc_info, msg=f'You are eligible to claim $MOVE: Now: {move_drop}, L2: {move_l2_drop}',
                        type_msg='success'
                    )
                else:
                    client.logger_msg(
                        *client.acc_info, msg=f'You are not eligible to claim $MOVE',
                        type_msg='error'
                    )
                    move_drop, move_l2_drop = '0', '0'
            else:
                raise SoftwareException(f"Can`t get nonce, response: {await response.text()}")

        except Exception as error:
            client.logger_msg(*client.acc_info, msg=f'Error: {error}')
            move_drop, move_l2_drop = 'ERROR', 'ERROR'

        return move_drop, move_l2_drop


    async def get_account_data(self, account_data, index):
        client = Client(account_data)
        account_name = account_data['account_name']
        account_address = client.address

        move_drop, move_l2_drop = await self.get_drop_amount(client)

        await client.session.close()

        full_data = {
            '#': index + 1,
            'Account Name': account_name,
            'Address': account_address,
            "MOVE": move_drop,
            "MOVE_L2": move_l2_drop,
        }

        return full_data

    async def check_wallets(self):
        cprint('✅  Processing wallets...')

        fields = ['#', 'Account Name', 'Address', 'MOVE', 'MOVE_L2']

        table = PrettyTable()
        table.field_names = [colored(field) for field in fields]

        accounts_data = []
        for account in zip(ACCOUNT_NAMES, PRIVATE_KEYS, PROXIES):
            if GeneralSettings.USE_PROXY:
                proxy = account[2]
            else:
                proxy = False

            account_data = {
                'account_name': account[0],
                'evm_private_key': account[1],
                'network': EthereumRPC,
                'proxy': proxy,
            }
            accounts_data.append(account_data)

        batch_size = 50
        batches = [accounts_data[i:i + batch_size] for i in range(0, len(accounts_data), batch_size)]

        wallets_data = []
        for batch_index, batch in enumerate(batches):
            cprint(f"🔁  Processing batch {batch_index + 1}/{len(batches)}...", 'light_yellow', attrs=["blink"])

            tasks = [self.get_account_data(account_data, index + batch_index * batch_size)
                     for index, account_data in enumerate(batch)]
            batch_wallets_data = await asyncio.gather(*tasks)

            wallets_data.extend(batch_wallets_data)

        numeric_columns = [field for field in fields if field not in ['#', 'Account Name', 'Address']]
        xlsx_data = pd.DataFrame(wallets_data)
        for column in numeric_columns:
            xlsx_data[column] = pd.to_numeric(xlsx_data[column], errors='coerce')
        directory = './data/accounts_stats/'
        if not os.path.exists(directory):
            os.makedirs(directory)
        excel_path = './data/accounts_stats/wallets_stats.xlsx'
        xlsx_data.to_excel(excel_path, index=False)

        workbook = load_workbook(excel_path)
        worksheet = workbook.active

        last_row = worksheet.max_row
        for col in range(4, worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            sum_formula = f'=SUM({col_letter}2:{col_letter}{last_row})'
            sum_cell = worksheet.cell(row=last_row + 1, column=col, value=sum_formula)
            sum_cell.font = Font(bold=True)

        total_label_cell = worksheet.cell(row=last_row + 1, column=3, value="Total")
        total_label_cell.font = Font(bold=True)

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
                cell.alignment = Alignment(horizontal='center', vertical='center')

            adjusted_width = (max_length + 4)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(excel_path)
        print()
        cprint('✅  Data successfully loaded to /data/accounts_stats/wallets_stats.xlsx (Excel format)\n',
               'light_green', attrs=["blink"])

        total_move, total_move_l2 = 0, 0

        for row in wallets_data:
            colored_row_data = []
            for field in fields:
                if field in ['#', 'Account Name', 'Address']:
                    colored_row_data.append(colored(row[field], 'light_green'))
                elif field in ['MOVE']:
                    total_move += float(row[field] if row[field] != 'ERROR' else 0)
                    colored_row_data.append(colored(row[field], 'light_magenta'))
                elif field in ['MOVE_L2']:
                    total_move_l2 += float(row[field] if row[field] != 'ERROR' else 0)
                    colored_row_data.append(colored(row[field], 'light_magenta'))
                else:
                    colored_row_data.append(colored(row[field], 'light_cyan'))
            table.add_row(colored_row_data, divider=True)

        print(table)

        cprint(f"TOTAL MOVE TO CLAIM ON MAINNET: {total_move}", 'light_magenta', attrs=['bold'])
        print()
        cprint(f"TOTAL MOVE TO CLAIM ON MOVEMENT: {total_move_l2}", 'light_magenta', attrs=['bold'])
